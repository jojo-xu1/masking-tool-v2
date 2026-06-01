from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .models import ReplacementRule, ReplacementTable, ReplacementTableError

REQUIRED_HEADERS = ("No", "検出語句", "置換提案")


def load_replacement_table(path: str | Path) -> ReplacementTable:
    table_path = Path(path)
    workbook = load_workbook(table_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ReplacementTableError("置換表が空です。")

    headers = tuple("" if value is None else str(value).strip() for value in rows[0])
    index = {name: headers.index(name) for name in headers if name}
    missing = [header for header in REQUIRED_HEADERS if header not in index]
    if missing:
        raise ReplacementTableError("`No`, `検出語句`, `置換提案` 列が必要です。")

    rules: list[ReplacementRule] = []
    seen: set[str] = set()
    for row_number, row in enumerate(rows[1:], start=2):
        no = _cell(row, index["No"]).strip()
        detected = _cell(row, index["検出語句"]).strip()
        replacement = _cell(row, index["置換提案"]).strip()
        if not detected and not replacement and not no:
            continue
        if not detected or not replacement:
            raise ReplacementTableError("置換表に空欄または重複した検出語句があります。")
        if detected in seen:
            raise ReplacementTableError("置換表に空欄または重複した検出語句があります。")
        seen.add(detected)
        rules.append(
            ReplacementRule(
                no=no,
                detected_phrase=detected,
                replacement_proposal=replacement,
                row_index=row_number,
            )
        )

    if not rules:
        raise ReplacementTableError("有効な置換ルールがありません。")

    return ReplacementTable(
        path=table_path,
        source_headers=headers,
        rules=tuple(sort_rules(rules)),
    )


def sort_rules(rules: Iterable[ReplacementRule]) -> list[ReplacementRule]:
    return sorted(
        rules,
        key=lambda rule: (-len(rule.detected_phrase), _sortable_no(rule.no), rule.row_index),
    )


def _cell(row: tuple[object, ...], index: int) -> str:
    if index >= len(row) or row[index] is None:
        return ""
    return str(row[index])


def _sortable_no(value: str) -> tuple[int, int | str]:
    try:
        return (0, int(value))
    except (TypeError, ValueError):
        return (1, value)
